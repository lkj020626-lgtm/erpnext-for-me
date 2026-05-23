import io

import frappe
from frappe import _

from starmake.master_data.import_export import (
    CUSTOMER_IMPORT_COLUMNS,
    ITEM_IMPORT_COLUMNS,
    SUPPLIER_IMPORT_COLUMNS,
)


@frappe.whitelist()
def download_template(doctype):
    """Generate and download an Excel import template for the given doctype."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        frappe.throw(_("openpyxl is required. Install with: pip install openpyxl"))

    columns_map = {
        "Item": ITEM_IMPORT_COLUMNS,
        "Customer": CUSTOMER_IMPORT_COLUMNS,
        "Supplier": SUPPLIER_IMPORT_COLUMNS,
    }
    columns = columns_map.get(doctype)
    if not columns:
        frappe.throw(_("Unsupported doctype: {0}").format(doctype))

    wb = Workbook()
    ws = wb.active
    ws.title = doctype

    header_font = Font(bold=True, size=11)
    required_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_def["label"]
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        if col_def["required"]:
            cell.fill = required_fill
        ws.column_dimensions[cell.column_letter].width = max(len(col_def["label"]) * 2.5, 12)

    ws.cell(row=2, column=1).value = "（黄色底色为必填字段，请从第2行开始填写数据）"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{doctype}_import_template.xlsx"
    frappe.response["filename"] = filename
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"
