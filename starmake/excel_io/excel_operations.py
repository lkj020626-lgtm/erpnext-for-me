import frappe
from frappe import _


INITIAL_STOCK_COLUMNS = [
    {"fieldname": "item_code", "label": "商品编码", "required": True},
    {"fieldname": "warehouse", "label": "仓库", "required": True},
    {"fieldname": "qty", "label": "数量", "required": True},
    {"fieldname": "rate", "label": "单价（估值）", "required": False},
    {"fieldname": "batch_no", "label": "批次号", "required": False},
]


@frappe.whitelist()
def get_initial_stock_template():
    """Return column definitions for initial stock import template."""
    return INITIAL_STOCK_COLUMNS


@frappe.whitelist()
def import_initial_stock(rows):
    """Import initial stock via Stock Reconciliation."""
    import json

    if isinstance(rows, str):
        rows = json.loads(rows)

    errors = []
    for idx, row in enumerate(rows, start=2):
        if not row.get("item_code"):
            errors.append({"row": idx, "error": "商品编码为空"})
            continue
        if not row.get("warehouse"):
            errors.append({"row": idx, "error": "仓库为空"})
            continue
        if not row.get("qty"):
            errors.append({"row": idx, "error": "数量为空"})
            continue
        if not frappe.db.exists("Item", row["item_code"]):
            errors.append({"row": idx, "error": f"商品 [{row['item_code']}] 不存在"})
        if not frappe.db.exists("Warehouse", row["warehouse"]):
            errors.append({"row": idx, "error": f"仓库 [{row['warehouse']}] 不存在"})

    if errors:
        return {"success": 0, "failed": len(errors), "errors": errors}

    try:
        recon = frappe.new_doc("Stock Reconciliation")
        recon.purpose = "Opening Stock"
        recon.posting_date = frappe.utils.today()

        for row in rows:
            recon.append(
                "items",
                {
                    "item_code": row["item_code"],
                    "warehouse": row["warehouse"],
                    "qty": float(row["qty"]),
                    "valuation_rate": float(row.get("rate") or 0) or None,
                    "batch_no": row.get("batch_no") or None,
                },
            )

        recon.insert(ignore_permissions=True)
        recon.submit()
        frappe.db.commit()

        return {
            "success": len(rows),
            "failed": 0,
            "errors": [],
            "reconciliation": recon.name,
        }
    except Exception as e:
        return {"success": 0, "failed": len(rows), "errors": [{"row": 0, "error": str(e)}]}


@frappe.whitelist()
def export_report_to_excel(report_name, filters=None):
    """Export any script report data to Excel."""
    import io
    import json

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        frappe.throw(_("openpyxl is required"))

    if isinstance(filters, str):
        filters = json.loads(filters)

    report = frappe.get_doc("Report", report_name)
    module_path = report.module.replace(" ", "_").lower()

    report_func = frappe.get_attr(
        f"starmake.reports.{report_name.lower().replace(' ', '_')}.{report_name.lower().replace(' ', '_')}.execute"
    )
    result = report_func(filters)
    columns = result[0]
    data = result[1]

    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]

    header_font = Font(bold=True)
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col.get("label", col.get("fieldname", ""))
        cell.font = header_font

    for row_idx, row in enumerate(data, start=2):
        for col_idx, col in enumerate(columns, start=1):
            fieldname = col.get("fieldname", "")
            value = row.get(fieldname, "") if isinstance(row, dict) else (row[col_idx - 1] if col_idx - 1 < len(row) else "")
            ws.cell(row=row_idx, column=col_idx).value = value

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_name}_{frappe.utils.today()}.xlsx"
    frappe.response["filename"] = filename
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"
